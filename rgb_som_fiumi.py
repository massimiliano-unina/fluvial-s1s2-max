# -*- coding: utf-8 -*-
"""
Created on Thu Jan 30 14:22:05 2020

@author: massi
"""
import os 
# import gdal
data_path = r"D:\fiumiunsupervised\drive-download-20210125T123716Z-001\out2\\"  #"D:\Albufera_2019_processed\subset_albufera\s1\\"
out_path = r"D:\fiumiunsupervised\drive-download-20210125T123716Z-001\out3\\"
# som_out_path = r"C:\Users\massi\Downloads\drive-download-20201201T153241Z-001\som\\"

otb_path = r"C:\Users\massi\Downloads\OTB-6.6.1-Win64\bin\\"

if not os.path.exists(out_path):
    os.makedirs(out_path)
# if not os.path.exists(som_out_path):
#     os.makedirs(som_out_path)

import numpy as np
    
dir_list = os.listdir(data_path)
dir_list.sort()
print(np.size(dir_list))
from openpyxl import load_workbook
workbook = load_workbook(filename=r"C:\Users\massi\OneDrive\Desktop\Incendi Boschivi\SQI_3_pH.xlsx")
print(workbook.sheetnames)
sheet_to_focus = 'SQI'

for s in range(len(workbook.sheetnames)):
    if workbook.sheetnames[s] == sheet_to_focus:
        break
workbook.active = s
sheet = workbook.active

for value in sheet.iter_rows(min_row=2,  min_col=2,max_col=5,values_only=True):
    print(value[2])
# print(sheet["B2:D4"].values)

for Num in range(np.size(dir_list)):
    print(dir_list[Num])

# for file in dir_list: 
#    if file.find("VV_Po_S1_pre_") != -1:
#         print(file)
#         file_inr1_pre = os.path.join(data_path, file )
#         name_ = 13
#         file_inr1_pre2 = os.path.join(out_path, file )
#         # file_inr2 = os.path.join(file_inr1[:len(data_path)], "B8" + file_inr1[len(data_path)+6:])
#         file_inr4= os.path.join(file_inr1_pre[:len(data_path)] ,"RF_Po_S1S2_" + file_inr1_pre[len(data_path)+name_:])

#         file_inr5= os.path.join(file_inr1_pre[:len(data_path)] ,"VH_Po_S1_" + file_inr1_pre[len(data_path)+name_:])
#         file_inr5_pre= os.path.join(file_inr1_pre[:len(data_path)] ,"VH_Po_S1_pre_" + file_inr1_pre[len(data_path)+name_:])
#         file_inr5_post= os.path.join(file_inr1_pre[:len(data_path)] ,"VH_Po_S1_post_" + file_inr1_pre[len(data_path)+name_:])
#         file_inr1= os.path.join(file_inr1_pre[:len(data_path)] ,"VV_Po_S1_" + file_inr1_pre[len(data_path)+name_:])
#         file_inr1_post= os.path.join(file_inr1_pre[:len(data_path)] ,"VV_Po_S1_post_" + file_inr1_pre[len(data_path)+name_:])

#         file_inr52= os.path.join(file_inr1_pre2[:len(out_path)] ,"VH_Po_S1_" + file_inr1_pre2[len(out_path)+name_:])
#         file_inr5_pre2= os.path.join(file_inr1_pre2[:len(out_path)] ,"VH_Po_S1_pre_" + file_inr1_pre2[len(out_path)+name_:])
#         file_inr5_post2= os.path.join(file_inr1_pre2[:len(out_path)] ,"VH_Po_S1_post_" + file_inr1_pre2[len(out_path)+name_:])
#         file_inr12= os.path.join(file_inr1_pre2[:len(out_path)] ,"VV_Po_S1_" + file_inr1_pre2[len(out_path)+name_:])
#         file_inr1_post2= os.path.join(file_inr1_pre2[:len(out_path)] ,"VV_Po_S1_post_" + file_inr1_pre2[len(out_path)+name_:])

#         # dataset = gdal.Open(file_inr4, gdal.GA_ReadOnly)
#         # rf = dataset.ReadAsArray()
#         # # gvv_0 = -10*np.log10(gvv_0)
#         # dataset = None

#         # rf3 = rf.astype('float32')
#         # file_inr42= os.path.join(file_inr1_pre2[:len(data_path)] ,"RF_Po_S1S2_" + file_inr1_pre2[len(data_path)+name_:])
#         # imsave(file_inr42, rf3)


#         conc = os.path.join(otb_path, "otbcli_Superimpose")
#         command = conc + " -inr " +  file_inr4 + " -inm " + file_inr1_pre + " -out " + file_inr1_pre2 + " -interpolator linear"
#         os.system(command)
#         command = conc + " -inr " +  file_inr4 + " -inm " + file_inr5  + " -out " + file_inr52 + " -interpolator linear"
#         os.system(command)
#         command = conc + " -inr " +  file_inr4 + " -inm " + file_inr5_pre  + " -out " + file_inr5_pre2 + " -interpolator linear"
#         os.system(command)
#         command = conc + " -inr " +  file_inr4 + " -inm " + file_inr5_post  + " -out " + file_inr5_post2 + " -interpolator linear"
#         os.system(command)
#         command = conc + " -inr " +  file_inr4 + " -inm " + file_inr1  + " -out " + file_inr12 + " -interpolator linear"
#         os.system(command)
#         command = conc + " -inr " +  file_inr4 + " -inm " + file_inr1_post  + " -out " + file_inr1_post2 + " -interpolator linear"
#         os.system(command)


# for file in dir_list: 
#    if file.find("VV_Osti_S1_pre_") != -1:
#         print(file)
#         file_inr1_pre = os.path.join(data_path, file )
#         name_ = 15
#         file_inr1_pre2 = os.path.join(out_path, file )
#         # file_inr2 = os.path.join(file_inr1[:len(data_path)], "B8" + file_inr1[len(data_path)+6:])
#         file_inr4= os.path.join(file_inr1_pre[:len(data_path)] ,"RF_Osti_S1S2_" + file_inr1_pre[len(data_path)+name_:])

#         file_inr5= os.path.join(file_inr1_pre[:len(data_path)] ,"VH_Osti_S1_" + file_inr1_pre[len(data_path)+name_:])
#         file_inr5_pre= os.path.join(file_inr1_pre[:len(data_path)] ,"VH_Osti_S1_pre_" + file_inr1_pre[len(data_path)+name_:])
#         file_inr5_post= os.path.join(file_inr1_pre[:len(data_path)] ,"VH_Osti_S1_post_" + file_inr1_pre[len(data_path)+name_:])
#         file_inr1= os.path.join(file_inr1_pre[:len(data_path)] ,"VV_Osti_S1_" + file_inr1_pre[len(data_path)+name_:])
#         file_inr1_post= os.path.join(file_inr1_pre[:len(data_path)] ,"VV_Osti_S1_post_" + file_inr1_pre[len(data_path)+name_:])

#         file_inr52= os.path.join(file_inr1_pre2[:len(out_path)] ,"VH_Osti_S1_" + file_inr1_pre2[len(out_path)+name_:])
#         file_inr5_pre2= os.path.join(file_inr1_pre2[:len(out_path)] ,"VH_Osti_S1_pre_" + file_inr1_pre2[len(out_path)+name_:])
#         file_inr5_post2= os.path.join(file_inr1_pre2[:len(out_path)] ,"VH_Osti_S1_post_" + file_inr1_pre2[len(out_path)+name_:])
#         file_inr12= os.path.join(file_inr1_pre2[:len(out_path)] ,"VV_Osti_S1_" + file_inr1_pre2[len(out_path)+name_:])
#         file_inr1_post2= os.path.join(file_inr1_pre2[:len(out_path)] ,"VV_Osti_S1_post_" + file_inr1_pre2[len(out_path)+name_:])

#         # dataset = gdal.Open(file_inr4, gdal.GA_ReadOnly)
#         # rf = dataset.ReadAsArray()
#         # # gvv_0 = -10*np.log10(gvv_0)
#         # dataset = None

#         # rf3 = rf.astype('float32')
#         # file_inr42= os.path.join(file_inr1_pre2[:len(data_path)] ,"RF_Osti_S1S2_" + file_inr1_pre2[len(data_path)+name_:])
#         # imsave(file_inr42, rf3)


#         conc = os.path.join(otb_path, "otbcli_Superimpose")
#         command = conc + " -inr " +  file_inr4 + " -inm " + file_inr1_pre + " -out " + file_inr1_pre2 + " -interpolator linear"
#         os.system(command)
#         command = conc + " -inr " +  file_inr4 + " -inm " + file_inr5  + " -out " + file_inr52 + " -interpolator linear"
#         os.system(command)
#         command = conc + " -inr " +  file_inr4 + " -inm " + file_inr5_pre  + " -out " + file_inr5_pre2 + " -interpolator linear"
#         os.system(command)
#         command = conc + " -inr " +  file_inr4 + " -inm " + file_inr5_post  + " -out " + file_inr5_post2 + " -interpolator linear"
#         os.system(command)
#         command = conc + " -inr " +  file_inr4 + " -inm " + file_inr1  + " -out " + file_inr12 + " -interpolator linear"
#         os.system(command)
#         command = conc + " -inr " +  file_inr4 + " -inm " + file_inr1_post  + " -out " + file_inr1_post2 + " -interpolator linear"
#         os.system(command)


# # data_path = r"D:\fiumiunsupervised\drive-download-20210125T123716Z-001\\"
# # out_path = r"D:\fiumiunsupervised\drive-download-20210125T123716Z-001\out2\\"
# # # som_out_path = r"C:\Users\massi\Downloads\drive-download-20201201T153241Z-001\som\\"

# # otb_path = r"C:\Users\massi\Downloads\OTB-6.6.1-Win64\bin\\"

# # if not os.path.exists(out_path):
# #     os.makedirs(out_path)
# # # if not os.path.exists(som_out_path):
# # #     os.makedirs(som_out_path)
    
    
# # dir_list = os.listdir(data_path)
# # dir_list.sort()

# # # for file in dir_list: 
# # #    if file.find("VV_Po_S1_pre_") != -1:
# # #         print(file)
# # #         file_inr1_pre = os.path.join(data_path, file )
# # #         name_ = 13
# # #         file_inr1_pre2 = os.path.join(out_path, file )
# # #         # file_inr2 = os.path.join(file_inr1[:len(data_path)], "B8" + file_inr1[len(data_path)+6:])
# # #         file_inr4= os.path.join(file_inr1_pre[:len(data_path)] ,"RF_Po_S1S2_" + file_inr1_pre[len(data_path)+name_:])

# # #         file_inr5= os.path.join(file_inr1_pre[:len(data_path)] ,"VH_Po_S1_" + file_inr1_pre[len(data_path)+name_:])
# # #         file_inr5_pre= os.path.join(file_inr1_pre[:len(data_path)] ,"VH_Po_S1_pre_" + file_inr1_pre[len(data_path)+name_:])
# # #         file_inr5_post= os.path.join(file_inr1_pre[:len(data_path)] ,"VH_Po_S1_post_" + file_inr1_pre[len(data_path)+name_:])
# # #         file_inr1= os.path.join(file_inr1_pre[:len(data_path)] ,"VV_Po_S1_" + file_inr1_pre[len(data_path)+name_:])
# # #         file_inr1_post= os.path.join(file_inr1_pre[:len(data_path)] ,"VV_Po_S1_post_" + file_inr1_pre[len(data_path)+name_:])

# # #         file_inr52= os.path.join(file_inr1_pre2[:len(out_path)] ,"VH_Po_S1_" + file_inr1_pre2[len(out_path)+name_:])
# # #         file_inr5_pre2= os.path.join(file_inr1_pre2[:len(out_path)] ,"VH_Po_S1_pre_" + file_inr1_pre2[len(out_path)+name_:])
# # #         file_inr5_post2= os.path.join(file_inr1_pre2[:len(out_path)] ,"VH_Po_S1_post_" + file_inr1_pre2[len(out_path)+name_:])
# # #         file_inr12= os.path.join(file_inr1_pre2[:len(out_path)] ,"VV_Po_S1_" + file_inr1_pre2[len(out_path)+name_:])
# # #         file_inr1_post2= os.path.join(file_inr1_pre2[:len(out_path)] ,"VV_Po_S1_post_" + file_inr1_pre2[len(out_path)+name_:])

        
# # #         # multi_file_inr1 = os.path.join(file_inr1_pre[:len(data_path)] ,"out\Multi_VVVH_Po_S1_" + file_inr1_pre[len(data_path)+name_:])
# # #         # dspk_file_inr1 = os.path.join(file_inr1_pre[:len(data_path)] ,"out\Dspk_Multi_VVVH_Po_S1_" + file_inr1_pre[len(data_path)+name_:])
# # #         # conc_vv = os.path.join(otb_path, "otbcli_ConcatenateImages")
# # #         # cmd_vv_vh = conc_vv + " -il " +  file_inr1_pre + " " +  file_inr1 + " "+  file_inr1_post + " " + file_inr5_pre + " " +  file_inr5 + " "+ file_inr5_post + " -out " + multi_file_inr1
# # #         # os.system(cmd_vv_vh)
# # #         # dspk_vv = os.path.join(otb_path, "otbcli_Despeckle")
# # #         # cmd_spk = dspk_vv +" -in " + multi_file_inr1 + " -filter gammamap -filter.gammamap.rad 3  -out " + dspk_file_inr1
# # #         # os.system(cmd_spk)
# # #         # single_vv = os.path.join(otb_path, "otbcli_BandMathX")
# # #         # vv_pre = single_vv +" -il " + dspk_file_inr1 + " -out " + file_inr1_pre2 + " -exp im1b1"
# # #         # vv_ = single_vv +" -il " + dspk_file_inr1 + " -out " + file_inr12 + " -exp im1b2"
# # #         # vv_post = single_vv +" -il " + dspk_file_inr1 + " -out " + file_inr1_post2 + " -exp im1b3"
# # #         # vh_pre = single_vv +" -il " + dspk_file_inr1 + " -out " + file_inr5_pre2 + " -exp im1b4"
# # #         # vh_ = single_vv +" -il " + dspk_file_inr1 + " -out " + file_inr52 + " -exp im1b5"
# # #         # vh_post = single_vv +" -il " + dspk_file_inr1 + " -out " + file_inr5_post2 + " -exp im1b6"
# # #         # os.system(vv_pre)
# # #         # os.system(vv_)
# # #         # os.system(vv_post)
# # #         # os.system(vh_pre)
# # #         # os.system(vh_)
# # #         # os.system(vh_post)
# # #         dspk_vv = os.path.join(otb_path, "otbcli_Despeckle")
# # #         cmd_spk = dspk_vv +" -in " + file_inr1_pre + " -filter gammamap -filter.gammamap.rad 3  -out " + file_inr1_pre2
# # #         os.system(cmd_spk)
# # #         cmd_spk = dspk_vv +" -in " + file_inr1 + " -filter gammamap -filter.gammamap.rad 3  -out " + file_inr12
# # #         os.system(cmd_spk)
# # #         cmd_spk = dspk_vv +" -in " + file_inr1_post + " -filter gammamap -filter.gammamap.rad 3  -out " + file_inr1_post2
# # #         os.system(cmd_spk)
# # #         cmd_spk = dspk_vv +" -in " + file_inr5_pre + " -filter gammamap -filter.gammamap.rad 3  -out " + file_inr5_pre2
# # #         os.system(cmd_spk)
# # #         cmd_spk = dspk_vv +" -in " + file_inr5 + " -filter gammamap -filter.gammamap.rad 3  -out " + file_inr52
# # #         os.system(cmd_spk)
# # #         cmd_spk = dspk_vv +" -in " + file_inr1_post + " -filter gammamap -filter.gammamap.rad 3  -out " + file_inr5_post2
# # #         os.system(cmd_spk)

# # for file in dir_list: 
# #    if file.find("VV_Osti_S1_pre_") != -1:
# #         print(file)
# #         file_inr1_pre = os.path.join(data_path, file )
# #         name_ = 15
# #         file_inr1_pre2 = os.path.join(out_path, file )
# #         # file_inr2 = os.path.join(file_inr1[:len(data_path)], "B8" + file_inr1[len(data_path)+6:])
# #         file_inr4= os.path.join(file_inr1_pre[:len(data_path)] ,"RF_Osti_S1S2_" + file_inr1_pre[len(data_path)+name_:])

# #         file_inr5= os.path.join(file_inr1_pre[:len(data_path)] ,"VH_Osti_S1_" + file_inr1_pre[len(data_path)+name_:])
# #         file_inr5_pre= os.path.join(file_inr1_pre[:len(data_path)] ,"VH_Osti_S1_pre_" + file_inr1_pre[len(data_path)+name_:])
# #         file_inr5_post= os.path.join(file_inr1_pre[:len(data_path)] ,"VH_Osti_S1_post_" + file_inr1_pre[len(data_path)+name_:])
# #         file_inr1= os.path.join(file_inr1_pre[:len(data_path)] ,"VV_Osti_S1_" + file_inr1_pre[len(data_path)+name_:])
# #         file_inr1_post= os.path.join(file_inr1_pre[:len(data_path)] ,"VV_Osti_S1_post_" + file_inr1_pre[len(data_path)+name_:])

# #         file_inr52= os.path.join(file_inr1_pre2[:len(out_path)] ,"VH_Osti_S1_" + file_inr1_pre2[len(out_path)+name_:])
# #         file_inr5_pre2= os.path.join(file_inr1_pre2[:len(out_path)] ,"VH_Osti_S1_pre_" + file_inr1_pre2[len(out_path)+name_:])
# #         file_inr5_post2= os.path.join(file_inr1_pre2[:len(out_path)] ,"VH_Osti_S1_post_" + file_inr1_pre2[len(out_path)+name_:])
# #         file_inr12= os.path.join(file_inr1_pre2[:len(out_path)] ,"VV_Osti_S1_" + file_inr1_pre2[len(out_path)+name_:])
# #         file_inr1_post2= os.path.join(file_inr1_pre2[:len(out_path)] ,"VV_Osti_S1_post_" + file_inr1_pre2[len(out_path)+name_:])

# #         dspk_vv = os.path.join(otb_path, "otbcli_Despeckle")
# #         cmd_spk = dspk_vv +" -in " + file_inr1_pre + " -filter gammamap -filter.gammamap.rad 3  -out " + file_inr1_pre2
# #         os.system(cmd_spk)
# #         cmd_spk = dspk_vv +" -in " + file_inr1 + " -filter gammamap -filter.gammamap.rad 3  -out " + file_inr12
# #         os.system(cmd_spk)
# #         cmd_spk = dspk_vv +" -in " + file_inr1_post + " -filter gammamap -filter.gammamap.rad 3  -out " + file_inr1_post2
# #         os.system(cmd_spk)
# #         cmd_spk = dspk_vv +" -in " + file_inr5_pre + " -filter gammamap -filter.gammamap.rad 3  -out " + file_inr5_pre2
# #         os.system(cmd_spk)
# #         cmd_spk = dspk_vv +" -in " + file_inr5 + " -filter gammamap -filter.gammamap.rad 3  -out " + file_inr52
# #         os.system(cmd_spk)
# #         cmd_spk = dspk_vv +" -in " + file_inr1_post + " -filter gammamap -filter.gammamap.rad 3  -out " + file_inr5_post2
# #         os.system(cmd_spk)



